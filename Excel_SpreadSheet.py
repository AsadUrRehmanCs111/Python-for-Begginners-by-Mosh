import openpyxl as xl
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
for row in range(2, sheet.max_row + 1):  # ignore cell 1 because it is heading
    cell = sheet.cell(row, 3)
    print(cell.value)
    corrected_price = cell.value * 0.9
    print(corrected_price)
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price
wb.save('transactions1.xlsx')

